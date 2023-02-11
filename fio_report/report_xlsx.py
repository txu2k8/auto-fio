#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
@author:TXU
@file:report_xlsx
@time:2022/12/04
@email:tao.xu2008@outlook.com
@description: 生成Excel报告
"""
import os
import string
from loguru import logger
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from fio_report.models import ExcelReportSettings, PerformanceResult


class ReportXlsx(object):
    """生成EXCEL报告"""
    def __init__(self, output_path, json_data, comments=None):
        if comments is None:
            comments = {}
        self.output_path = output_path
        self.json_data = json_data
        self.comments = comments
        self.settings = ExcelReportSettings()

        self.title_row_count = len(self.settings.data_column_title)
        self.row_count = len(self.settings.data_column_title)  # 数据行数从表头行后开始
        self.col_max = string.ascii_uppercase[len(self.settings.data_column_title[0])-1]  # 最大的列名
        self.write_idx = self.settings.data_column_write_idx  # 写结果 idx
        self.read_idx = self.settings.data_column_write_idx + len(PerformanceResult().dict().keys())  # 读结果 idx

        self.wb = Workbook(write_only=False)
        # 创建4张表
        self.data_ws = self.wb.create_sheet(self.settings.data_sheet_title, self.settings.data_sheet_index)
        self.chart_ws = self.wb.create_sheet(self.settings.chart_sheet_title, self.settings.chart_sheet_index)
        self.desc_ws = self.wb.create_sheet(self.settings.desc_sheet_title, self.settings.desc_sheet_index)
        self.env_ws = self.wb.create_sheet(self.settings.env_sheet_title, self.settings.env_sheet_index)

    @staticmethod
    def get_item_index(items, key):
        """
        获取 key在items中的列位置
        :param items:
        :param key:
        :return:
        """
        for idx, item in enumerate(items):
            if item == key:
                return idx+1
        raise Exception("{}中没找到:{}".format(items, key))

    @staticmethod
    def reset_col_width(ws: Worksheet):
        """
        重置表格的列宽
        """
        for idx, col in enumerate(ws.columns):
            letter = get_column_letter(idx + 1)  # 列字母
            collen = max([len(str(c.value).encode()) for c in col])  # 获取这一列长度的最大值
            ws.column_dimensions[letter].width = collen  # collen * 1.2 + 4 也就是列宽为最大长度*1.2+4 可以自己调整

    def reset_data_sheet_style(self):
        # 设置字体
        title_font = Font(name='微软雅黑', size=12, b=True)
        data_font = Font(name='微软雅黑', size=11, b=False)
        # 设置文本对齐
        ali = Alignment(horizontal='center', vertical='center')
        """
        horizontal:水平对齐('centerContinuous', 'general', 'distributed', 'left', 'fill', 'center', 'justify', 'right')
        vertical:垂直对齐（'distributed', 'top', 'center', 'justify', 'bottom'）
        """
        # 设置图案填充，颜色一般使用十六进制RGB，'solid'是图案填充类型，详细可查阅文档
        fill = PatternFill('solid', fgColor='90B248')
        fill_write = PatternFill('solid', fgColor='BCD08A')
        fill_read = PatternFill('solid', fgColor='E6EDD4')
        # 设置边框
        side = Side(style='thin', color='000000')  # 设置边框样式
        border = Border(top=side, bottom=side, left=side, right=side)

        data = self.data_ws[f"A1:{self.col_max}{self.row_count}"]
        for idx_r, row in enumerate(data):
            for idx_c, item in enumerate(row):
                item.border = border
                item.font = data_font
                if idx_r < self.title_row_count:
                    # 表头
                    item.alignment = ali
                    item.font = title_font
                    if idx_c < self.write_idx:
                        item.fill = fill
                    elif idx_c < self.read_idx:
                        item.fill = fill_write
                    else:
                        item.fill = fill_read
                if idx_c < self.write_idx:
                    # 描述列 居中
                    item.alignment = ali
                    if idx_c == 0 and idx_r > 0:
                        # 第一列除表头外，左对齐
                        item.alignment = Alignment(horizontal='left', vertical='center')

    def insert_to_data_sheet(self):
        """
        data数据表
        """
        # 插入表头
        for column_title in self.settings.data_column_title:
            self.data_ws.append(column_title)
        if self.title_row_count > 1:
            for c in string.ascii_uppercase[:self.write_idx]:
                self.data_ws.merge_cells(f'{c}1:{c}2')
            # self.data_ws.merge_cells('J1:L1')
            # self.data_ws.merge_cells('M1:O1')

        # 插入数据
        for bs_data in self.json_data:
            for data in bs_data['data']:
                row = [
                    # 测试 重要参数
                    data['name'],
                    data['client_num'],
                    data['ioengine'],
                    data['runtime'],
                    data['direct'],
                    data['iodepth'],
                    data['numjobs'],
                    data['filesize'],
                    data['bs'],
                    data['rw'],
                    # 写 结果
                    0,
                    0,
                    0,
                    # 读 结果
                    0,
                    0,
                    0,
                ]
                for result in data["result"]:
                    idx_start = self.write_idx if result["type"] == "write" else self.read_idx
                    row[idx_start+0] = result['bw_mb']
                    row[idx_start+1] = result['iops']
                    row[idx_start+2] = result['clat_ms']  # lat_ms
                logger.debug(row)
                self.data_ws.append(row)
                self.row_count += 1
        self.data_ws.auto_filter.ref = f"A{self.title_row_count}:I{self.row_count}"
        self.data_ws.auto_filter.add_filter_column(0, [])
        # self.data_ws.auto_filter.add_sort_condition(f"F3:F{self.row_count}")
        # self.data_ws.auto_filter.add_sort_condition(f"G3:G{self.row_count}")
        # self.data_ws.auto_filter.add_sort_condition(f"H3:H{self.row_count}")
        # 重置列宽
        self.reset_col_width(self.data_ws)
        # 重置格式
        self.reset_data_sheet_style()

    def insert_to_desc_sheet(self):
        """
        结果数据后，写入备注信息
        :return:
        """
        # 插入表头
        self.desc_ws.append(self.settings.desc_column_title)
        # 插入数据
        idx = 1
        for k, v in self.comments.items():
            idx += 1
            self.desc_ws[f"A{idx}"] = k
            self.desc_ws[f"B{idx}"] = v
            # self.desc_ws.merge_cells(f'B{idx}:K{idx}')  # 单独表格，不需要合并
        # 重置列宽
        self.reset_col_width(self.desc_ws)

    def bar_chart(self, key, x_title=None, y_title=None):
        """
        创建bar chart
        :param key:
        :param x_title:
        :param y_title:
        :return:
        """
        chart = BarChart()
        chart.type = "bar"
        chart.style = 13
        # chart.overlap = 100
        chart.height = 6+(self.row_count*1.2)
        chart.legend.position = "b"  # 下方 显示图例
        chart.title = "FIO性能对比 - {}".format(key.split("-")[0].upper())
        chart.y_axis.title = y_title or 'Test number'
        chart.x_axis.title = x_title or 'Test Job Name'
        col_index = self.get_item_index(self.settings.data_column_title[0], key+"-write")  # 1, key
        data1 = Reference(self.data_ws, min_row=1, max_row=self.row_count, min_col=col_index, max_col=col_index)
        data2 = Reference(self.data_ws, min_row=1, max_row=self.row_count, min_col=col_index+3, max_col=col_index+3)
        cats = Reference(self.data_ws, min_row=2, max_row=self.row_count, min_col=1)
        chart.add_data(data1, titles_from_data=True)
        chart.add_data(data2, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        # 显示数据标签
        for s in chart.series:
            s.dLbls = DataLabelList()
            s.dLbls.showVal = True

        return chart

    def bar_chart_3d(self, key):
        col_index = self.get_item_index(self.settings.data_column_title[0], key + "-write")  # 1, key
        col_index2 = self.get_item_index(self.settings.data_column_title[0], "iodepth")  # 1, key
        data = Reference(self.data_ws, min_row=1, max_row=self.row_count, min_col=col_index, max_col=col_index)
        data2 = Reference(self.data_ws, min_row=1, max_row=self.row_count, min_col=col_index2, max_col=col_index2)
        cats = Reference(self.data_ws, min_row=2, max_row=self.row_count, min_col=1)
        chart = BarChart3D()
        chart.title = "3D Bar Chart"
        chart.add_data(data=data, titles_from_data=True)
        chart.add_data(data=data2, titles_from_data=True)
        chart.set_categories(cats)
        return chart

    def bar_chart_bw(self):
        return self.bar_chart("bw", y_title="Test number(MiB)")

    def bar_chart_iops(self):
        return self.bar_chart("iops")

    def bar_chart_lat(self):
        return self.bar_chart("latency", y_title="Test number(ms)")

    def insert_to_bar_chart_sheet(self):
        self.chart_ws.add_chart(self.bar_chart_bw(), "A1")
        self.chart_ws.add_chart(self.bar_chart_iops(), "I1")
        self.chart_ws.add_chart(self.bar_chart_lat(), "Q1")

    def create_xlsx_file(self):
        # 写入数据统计表
        self.insert_to_data_sheet()
        # 写入描述表
        self.insert_to_desc_sheet()
        # 写入性能对比图
        self.insert_to_bar_chart_sheet()
        # 保存
        self.wb.save(os.path.join(os.path.abspath(self.output_path), "report.xlsx"))


if __name__ == '__main__':
    pass
